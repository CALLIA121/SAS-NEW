import sqlite3 as sq
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from settings import connect, cursor, DBlist, Say
import openpyxl

def GetRaw(execute):
    cursor.execute(execute)
    return cursor.fetchall()

def execute(execute):
    cursor.execute(execute)
    connect.commit()

def WriteData(DB: int,
           st: str,
           value,
           qvest = None) -> None:
    '''
    :param st: столбец для записи.
    :param DB: 1 - Users 2 - Data.
    :param qvest: условия для записи, изначально ID, напишите "!" в начале, для передачи SQ3 условия.
    :param value: данные для записи.'''
    DB = DBlist[DB]
    if qvest is None:
        if Say:
            print(f'Write INTO `{DB}` {st} VALUES "{value}"')
        cursor.execute(f'INSERT INTO `{DB}` ({st}) VALUES ({value})')
        return None
    qvest = str(qvest)
    if qvest[0] != "!":
        qvest = int(qvest)
        if Say:
            print(f'Write INTO `{DB}` {st} VALUES "{value}", WHERE `ID` = {qvest}')
        cursor.execute(f'''SELECT * FROM `{DB}` WHERE `ID` = "{qvest}"''')
        result = cursor.fetchone()
        if result is None:
            cursor.execute(f'INSERT INTO `{DB}` {st} VALUES (?)', value)
        else:
            cursor.execute(f'''UPDATE `{DB}` SET {st} = ? WHERE `ID` = ?''', (value, qvest))
    else:
        qvest = qvest[1:]
        cursor.execute(f'''SELECT * FROM `{DB}` WHERE "{qvest}"''')
        result = cursor.fetchone()
        if result is None:
            cursor.execute(f'INSERT INTO `{DB}` {st} VALUES (?)', (value, ))
        else:
            cursor.execute(f'''UPDATE `{DB}` SET `{st}` = ? WHERE  "{qvest}"''', (value, ))
    connect.commit()
        
def GetData(DB, st, qvest="!1=1"):
    '''
    :param St: столбец для чтения.
    :param DB: 1 - Users 2 - Data.
    :param Qest: условия для чтения, изначально ID, напишите "!" в начале, для передачи SQ3 условия.'''
    DB = DBlist[DB]
    if Say:
        print(f'''SELECT `{st}` FROM {DB} WHERE `ID` = "{qvest}"''')
    qvest = str(qvest)
    if qvest[0] != "!":
        cursor.execute(f'''SELECT `{st}` FROM {DB} WHERE `ID` = "{qvest}"''')
        result = cursor.fetchall()
        if Say:
            print("result", result)
        if result is None or result == [None]:
            return []
        else:
            return result
    elif st == "*":
        cursor.execute(f'''SELECT * FROM {DB}''')
    elif qvest[0] == "!":
        qvest = qvest[1:]
        cursor.execute(f'''SELECT `{st}` FROM {DB} WHERE {qvest}''')
        result = cursor.fetchall()
        if Say:
            print(f'''SELECT `{st}` FROM {DB} WHERE {qvest}''')
            print("result", result)
        if result is None or result == [None]:
            return []
        else:
            return result



def get_users():
    try:
        cursor.execute("SELECT ID, Name, PhotoPath, FaceEncoding FROM Users")
        users = cursor.fetchall()
        user_list = [{"ID": user[0], "Name": user[1], "PhotoPath": user[2], "FaceEncoding": user[3]} for user in users]
        print(f"Загружены пользователей: {len(user_list)}")
        return user_list
    except sq.Error as e:
        print(f"Ошибка при получении пользователей: {e}")
        return []

def get_user_by_id(user_id):
    try:
        cursor.execute("SELECT ID, Name, PhotoPath FROM Users WHERE ID = ?", (user_id,))
        user = cursor.fetchone()
        if user:
            user_data = {"ID": user[0], "Name": user[1], "PhotoPath": user[2]}
            print(f"Данные пользователя: {user_data}")
            return user_data
        else:
            print(f"Пользователь с ID {user_id} не найден")
            return None
    except sq.Error as e:
        print(f"Ошибка при получении пользователя с ID {user_id}: {e}")
        return None

def update_user_code(ID, newCode):
    try:
        cursor.execute("UPDATE Users SET FaceEncoding = ? WHERE ID = ?", (newCode, ID))
        connect.commit()
        print(f"Код лица с ID {ID} обновлен")
    except sq.Error as e:
        print(f"Ошибка при кода лица с ID {ID}: {e}")

def update_user_data(user_id, new_value):
    try:
        cursor.execute("UPDATE Users SET Name = ? WHERE ID = ?", (new_value, user_id))
        connect.commit()
        print(f"Пользователь с ID {user_id} обновлен до {new_value}")
    except sq.Error as e:
        print(f"Ошибка при обновлении пользователя с ID {user_id}: {e}")

def delete_user(user_id):
    try:
        cursor.execute("DELETE FROM Users WHERE ID = ?", (user_id,))
        connect.commit()
        print(f"Пользователь с ID {user_id} удален")
    except sq.Error as e:
        print(f"Ошибка при удалении пользователя с ID {user_id}: {e}")

def add_user(name, photo_path):
    try:
        cursor.execute("INSERT INTO Users (Name, PhotoPath) VALUES (?, ?)", (name, photo_path))
        connect.commit()
        print(f"Добавлен пользователь {name}")
    except sq.Error as e:
        print(f"Ошибка при добавлении пользователя: {e}")

def commit_changes():
    try:
        connect.commit()
        print("Изменения сохранены")
    except sq.Error as e:
        print(f"Ошибка при сохранении изменений: {e}")


def dataToExcel(Data, path, mode='replace'):
    """
    Записывает данные в Excel файл.

    :param Data: Данные для записи, в формате словаря массивов.
    :param path: Путь к Excel файлу.
    :param mode: Режим записи: 'over' для добавления данных, 'replace' для замены данных.
    :return: Путь к итоговому файлу.
    """
    # Открытие существующего файла Excel
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active  # Выбор активного листа
    
    # Если режим 'replace', очищаем все данные на листе
    if mode == 'replace':
        sheet.delete_rows(1, sheet.max_row)

    # Если лист был очищен или он пуст, добавляем заголовки
    if sheet.max_row == 1:
        headers = list(Data.keys())
        sheet.append(headers)

    # Определение стартовой строки для записи данных
    start_row = sheet.max_row + 1 if mode == 'over' else 1
    thin = Side(border_style="thin", color="000000")
    
    # Заполнение строк данными
    for i in range(len(next(iter(Data.values())))):
        row = [Data[key][i] for key in Data.keys()]
        sheet.append(row)

    ceils = sheet['A:D']
    for line in ceils:
        for ceil in line:
            ceil.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Сохранение изменений в Excel файл
    workbook.save(path)
    
    return path
